import os
import time
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

start_time = time.time()

main_folder = r"C:\Users\ymohdzaifullizan\OneDrive - Dyson\Year 2 rotation - E&O\Consolidate Exposure\Test setup 2\Sample file 2"
output_folder = r"C:\Users\ymohdzaifullizan\OneDrive - Dyson\Year 2 rotation - E&O\Consolidate Exposure\Test setup 2\Python outputs"
output_file = os.path.join(output_folder, summary_filename)
summary_filename = "master_summary.xlsx"
headers = [
    'Initial Claim submission Date','CR Number','CR Description','EOP Strategy','CM','EOP Declaration Timing','Last Time Build','Dyson PIC',
    'Product Category','Project','Model', 'Initial Submission Currency','Initial Submission','Claim Received (RM)','Claim Accepted (RM)','Claim value pending SAF/PR approval (RM)',
    'Claim Avoided (RM)','Claim in Progress (RM)','WIP (RM/USD)','Remark/Current Status','One Time Settlement','Claim Status','Finance Status',
    'CM Claim No (Commercial Title)','PR Number','PO Number','GR Status','GR Amount','Accrued/GR Amt','Provision','Check'
]
claim_fields = [
    'Claim Received (RM)', 'Claim Accepted (RM)', 'Claim value pending SAF/PR approval (RM)', 
    'Claim Avoided (RM)', 'Claim in Progress (RM)'
]

def parse_model_project(val):
    if isinstance(val, str):
        parts = val.strip().split()
        if len(parts) == 2:
            return parts[0], parts[1]
    return None, None

def extract_exposure_metadata(df):
    mapping = {}
    for i in range(6, 16):  # Rows 7-16 (index 6–15)
        cells = [
            (str(df.iloc[i, 0]).strip().lower() if pd.notnull(df.iloc[i, 0]) else "", df.iloc[i, 2]),
            (str(df.iloc[i, 4]).strip().lower() if pd.notnull(df.iloc[i, 4]) else "", df.iloc[i, 5])
        ]
        for k, v in cells:
            if not k: continue
            if "eop declare date" in k:
                mapping["EOP Declaration Timing"] = v
            elif "initial submission date" in k:
                mapping["Initial Claim submission Date"] = v
            elif "ltb week" in k:
                mapping["Last Time Build"] = v
            elif "initial submission value" in k:
                mapping["Initial Submission"] = v
            elif "contract manufacturing" in k:
                mapping["CM"] = v
            elif "currency" in k:
                mapping["Currency"] = v
            elif "category" in k:
                mapping["Product Category"] = v
            elif "exchange rate" in k:
                mapping["Exchange rate to MYR"] = v
            elif "model name" in k:
                model, project = parse_model_project(v)
                mapping["Model"] = model
                mapping["Project"] = project
            elif "cm claim no" in k:
                mapping["CM Claim No (Commercial Title)"] = v
            elif "dyson pic" in k:
                mapping["Dyson PIC"] = v
            elif "claim status" in k:
                mapping["Claim Status"] = v
            elif "cm pic" in k:
                mapping["CM PIC"] = v
            elif "pr number" in k:
                mapping["PR Number"] = v
            elif "remarks" in k:
                mapping["CR Number"] = v
            elif "po number" in k:
                mapping["PO Number"] = v
            elif "cr description" in k:
                mapping["CR Description"] = v
            elif "gr amount" in k:
                mapping["GR Amount"] = v
            elif "eop stratergy" in k or "eop strategy" in k:
                mapping["EOP Strategy"] = v
            elif "ranging out" in k:
                mapping["Ranging Out"] = v
            elif "cr-" in str(v).lower():
                mapping['CR Number'] = v
    return mapping

def extract_claim_fields(claim_df):
    claim_map = {}
    submission_currency = claim_df.iloc[37,1] #B38 == index 37; this is the currency for all claim values
    claim_map["Initial Submission Currency"] = submission_currency if pd.notnull(submission_currency) else ""
    # C&D == 2,3; rows 38-43 == index 37-42
    for i in range(37, 43):
        k = str(claim_df.iloc[i,2]).strip().lower() if pd.notnull(claim_df.iloc[i,2]) else ""
        v = claim_df.iloc[i,3]
        # Flexible matching for all five fields:
        if "claim received" in k:
            claim_map["Claim Received (RM)"] = v
        elif "claim accepted" in k:
            claim_map["Claim Accepted (RM)"] = v
        elif "claim value" in k and "saf" in k:
            claim_map["Claim value pending SAF/PR approval (RM)"] = v
        elif "claim avoided" in k:
            claim_map["Claim Avoided (RM)"] = v
        elif "claim in progress" in k:
            claim_map["Claim in Progress (RM)"] = v
    return claim_map


# ... (rest of your code stays as before)
output_rows = []
ranging_out_lookup = defaultdict(list)  # year:str -> [row, row, row, ...]

for filename in os.listdir(main_folder):
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        filepath = os.path.join(main_folder, filename)
        try:
            xls = pd.ExcelFile(filepath)
            df = pd.read_excel(xls, sheet_name="Appendix 2", header=None)
            meta = extract_exposure_metadata(df)
            # Get "Ranging Out" year (add fallback if missing)
            ranging_out = str(meta.get("Ranging Out", "NoYear")).strip()
            # Now extract claim fields from "Mitigation Summary Tracker"
            df2 = pd.read_excel(xls, sheet_name="Mitigation Summary Tracker", header=None)
            claim_meta = extract_claim_fields(df2)
            meta.update(claim_meta)
            row = [meta.get(h, "") for h in headers]
            ranging_out_lookup[ranging_out].append(row)
            print(f"Processed: {filename} (Ranging Out: {ranging_out})")
        except Exception as e:
            print(f"Error processing {filename}: {e}")

# Now write each year's rows to a separate sheet
with pd.ExcelWriter(os.path.join(main_folder, summary_filename), engine='openpyxl') as writer:
    for year, rows in ranging_out_lookup.items():
        sheet_name = f"Masterfile {year}"
        df_out = pd.DataFrame(rows, columns=headers)
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Summary saved as {summary_filename} in {main_folder}")

def format_initial_submission_column_by_currency(file_path, sheet_names):
    wb = load_workbook(file_path)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        # Find the column indices for your headers
        headers_row = 1
        initial_sub_col = None
        currency_col = None
        # Find headers
        for idx, cell in enumerate(ws[headers_row], start=1):
            if cell.value == "Initial Submission":
                initial_sub_col = get_column_letter(idx)
            if cell.value == "Initial Submission Currency":
                currency_col = get_column_letter(idx)
        if not initial_sub_col or not currency_col:
            print(f"Couldn't find required columns in {sheet_name}, skipping.")
            continue
        for row in range(2, ws.max_row + 1):
            currency = ws[f"{currency_col}{row}"].value
            cell = ws[f"{initial_sub_col}{row}"]
            if currency == "MYR" or currency == "RM":
                cell.number_format = '_("RM"* #,##0.00_);_("RM"* (#,##0.00);_("RM"* "-"??_);_(@_)'
            elif currency == "USD":
                cell.number_format = '_("$"* #,##0.00_);_("$"* (#,##0.00);_("$"* "-"??_);_(@_)'
            elif currency == "EUR":
                cell.number_format = _('_("€"* #,##0.00_);_("€"* (#,##0.00);_("€"* "-"??_);_(@_)')
            elif currency in ["GBP", "£"]:
                cell.number_format = '_("£"* #,##0.00_);_("£"* (#,##0.00);_("£"* "-"??_);_(@_)'
            elif currency in ["CNY", "RMB", "¥"]:
                cell.number_format = '_("¥"* #,##0.00_);_("¥"* (#,##0.00);_("¥"* "-"??_);_(@_)'
            # Add more currency logic as needed.
    wb.save(file_path)
    print("Formatting applied!")

# Usage after saving your summary file:
sheet_names = [f"Masterfile {year}" for year in ranging_out_lookup.keys()]
format_initial_submission_column_by_currency(
    output_file,
    sheet_names
)
end_time = time.time()
elapsed_time = end_time - start_time
print(f"Elapsed time: {elapsed_time:.2f} seconds")